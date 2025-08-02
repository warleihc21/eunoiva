import base64
import json
import re
import pandas as pd
import csv
import openpyxl
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, JsonResponse
import requests
from core import settings
from .models import Convidados, ImagemGaleria, ImagemNoivos, MensagemAosNoivos, MensagemSobreNoivoNoiva, Presentes, MensagemPersonalizada, ProdutoBase
from django.contrib.auth.decorators import login_required # type: ignore
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.shortcuts import render
import logging
from django.db.models import ProtectedError
from django.http import JsonResponse
from .models import Perfil
import webbrowser
from urllib.parse import quote
from time import sleep

from io import StringIO
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime
from django.core.files.storage import default_storage
import os
from io import BytesIO
from django.core.files.base import ContentFile
from urllib.parse import unquote
from django.views.decorators.http import require_POST



if os.environ.get('DISPLAY', None):  # Verifica se a variável de ambiente DISPLAY está configurada
    import pyautogui
else:
    print("Ambiente sem display gráfico. pyautogui não será usado.")




@login_required(login_url='/auth/logar/')
def home(request):
    if request.method == "GET":
        presentes = Presentes.objects.filter(user=request.user)
        nao_reservado = presentes.filter(reservado=False).count()
        reservado = presentes.filter(reservado=True).count()

        presentes_reservados = presentes.filter(reservado=True)
        total_reservado = sum(presente.preco for presente in presentes_reservados)

        perfil = Perfil.objects.get(user=request.user)
        mensagens = MensagemAosNoivos.objects.filter(user=request.user).order_by('-data_envio')
        todas_imagens = ImagemGaleria.objects.filter(perfil=perfil)
        
        mensagem_noiva = perfil.mensagens.filter(tipo='noiva').first()
        mensagem_noivo = perfil.mensagens.filter(tipo='noivo').first()

        imagem_noiva = perfil.fotosnoivos.filter(tipo='noiva').first()
        imagem_noivo = perfil.fotosnoivos.filter(tipo='noivo').first()

        timestamp = int(datetime.now().timestamp())

        # Verificar se o usuário já tem presentes
        tem_presentes = presentes.exists()
        # Verificar quantos produtos base existem
        tem_lista_padrao = ProdutoBase.objects.filter(ativo=True).exists()

        return render(request, 'home.html', {
            'presentes': presentes,
            'data': [nao_reservado, reservado],
            'presentes_reservados': presentes_reservados,
            'total_reservado': total_reservado,
            'nome_primeiro_conjuge': perfil.nome_primeiro_conjuge,
            'nome_segundo_conjuge': perfil.nome_segundo_conjuge,
            'data_casamento': perfil.data_casamento,
            'horario_casamento': perfil.horario_casamento,
            'imagem': perfil.imagem,
            'perfil': perfil,
            'mensagens': mensagens,
            'todas_imagens': todas_imagens,
            'timestamp': timestamp,
            'rua': perfil.rua,
            'numero': perfil.numero,
            'bairro': perfil.bairro,
            'municipio': perfil.municipio,
            'estado': perfil.estado,
            'pais': perfil.pais,
            'cep': perfil.cep,
            'mensagem_noiva': mensagem_noiva.mensagem if mensagem_noiva else '',
            'mensagem_noivo': mensagem_noivo.mensagem if mensagem_noivo else '',
            'imagem_noiva': imagem_noiva.imagem if imagem_noiva else '',
            'imagem_noivo': imagem_noivo.imagem if imagem_noivo else '',
            'tem_presentes': tem_presentes,
            'tem_lista_padrao': tem_lista_padrao,
        })

    elif request.method == "POST":
        presente_id = request.POST.get('presente_id')
        nome_presente = request.POST.get('nome_presente')
        foto = request.FILES.get('foto')
        preco = request.POST.get('preco')
        link_sugestao_compra = request.POST.get('link_sugestao_compra')
        link_cobranca = request.POST.get('link_cobranca')

        if ',' in preco:
            preco = preco.replace(',', '.')
        preco = float(preco)
        importancia = int(request.POST.get('importancia', 0) or 0)

        if presente_id:
            presente = Presentes.objects.get(id=presente_id, user=request.user)
            presente.nome_presente = nome_presente
            presente.preco = preco
            presente.importancia = importancia
            presente.link_sugestao_compra = link_sugestao_compra
            presente.link_cobranca = link_cobranca

            if foto:  
                presente.foto = foto

            presente.save()
        else:
            # Antes de criar um novo, verifica se já foi salvo pela API
            if not Presentes.objects.filter(nome_presente=nome_presente, user=request.user).exists():
                Presentes.objects.create(
                    user=request.user,
                    nome_presente=nome_presente,
                    foto=foto,
                    preco=preco,
                    importancia=importancia,
                    link_sugestao_compra=link_sugestao_compra,
                    link_cobranca=link_cobranca,
                )

    return redirect('home')




@login_required(login_url='/auth/logar/')
def substituir_imagem(request):
    if request.method == "POST":
        imagem_id = request.POST.get("imagem_id")
        nova_imagem_id = request.POST.get("nova_imagem_id")

        imagem_atual = get_object_or_404(ImagemGaleria, id=imagem_id)
        nova_imagem = get_object_or_404(ImagemGaleria, id=nova_imagem_id)

        # Atualiza a imagem no perfil
        imagem_atual.imagem = nova_imagem.imagem
        imagem_atual.save()

        return JsonResponse({"sucesso": True})

    return JsonResponse({"sucesso": False})


@login_required(login_url='/auth/logar/')
@require_POST
def excluir_imagem_galeria(request):
    try:
        imagem_id = request.POST.get('imagem_id')
        imagem = ImagemGaleria.objects.get(id=imagem_id, perfil=request.user.perfil)
        imagem.delete()
    except Exception as e:
        # Aqui você pode adicionar mensagens de erro com django.contrib.messages se quiser
        pass

    return redirect(request.META.get('HTTP_REFERER', '/'))  # Redireciona para a página anterior

@login_required(login_url='/auth/logar/')
def substituir_imagem_noivos(request):
    if request.method == "POST":
        perfil = Perfil.objects.get(user=request.user)

        # Verificando se as imagens da noiva e do noivo foram enviadas
        imagem_noiva = request.FILES.get('imagem_noiva')
        imagem_noivo = request.FILES.get('imagem_noivo')

        # Atualizando imagem da noiva
        if imagem_noiva:
            imagem_noiva_obj = ImagemNoivos.objects.filter(perfil=perfil, tipo='noiva').first()
            if imagem_noiva_obj:
                imagem_noiva_obj.imagem = imagem_noiva
                imagem_noiva_obj.save()
            else:
                ImagemNoivos.objects.create(perfil=perfil, tipo='noiva', imagem=imagem_noiva)

        # Atualizando imagem do noivo
        if imagem_noivo:
            imagem_noivo_obj = ImagemNoivos.objects.filter(perfil=perfil, tipo='noivo').first()
            if imagem_noivo_obj:
                imagem_noivo_obj.imagem = imagem_noivo
                imagem_noivo_obj.save()
            else:
                ImagemNoivos.objects.create(perfil=perfil, tipo='noivo', imagem=imagem_noivo)

        return redirect('home')
    
    return redirect('home')


@login_required(login_url='/auth/logar/')
def editar_mensagem(request):
    if request.method == "POST":
        perfil = Perfil.objects.get(user=request.user)

        # Obter os valores do POST
        mensagem_noiva_texto = request.POST.get('mensagem_noiva', '').strip()
        mensagem_noivo_texto = request.POST.get('mensagem_noivo', '').strip()

        # Atualizar mensagem da noiva
        if mensagem_noiva_texto:
            mensagem_noiva = perfil.mensagens.filter(tipo='noiva').first()
            if mensagem_noiva:
                mensagem_noiva.mensagem = mensagem_noiva_texto
                mensagem_noiva.save()
            else:
                MensagemSobreNoivoNoiva.objects.create(perfil=perfil, tipo='noiva', mensagem=mensagem_noiva_texto)

        # Atualizar mensagem do noivo
        if mensagem_noivo_texto:
            mensagem_noivo = perfil.mensagens.filter(tipo='noivo').first()
            if mensagem_noivo:
                mensagem_noivo.mensagem = mensagem_noivo_texto
                mensagem_noivo.save()
            else:
                MensagemSobreNoivoNoiva.objects.create(perfil=perfil, tipo='noivo', mensagem=mensagem_noivo_texto)

        return redirect('home')


   




def lista_convidados(request):
    if request.method == 'GET':
        convidados = Convidados.objects.filter(user=request.user)
        nao_confirmados = convidados.filter(status='AC')
        
        mensagem_obj = MensagemPersonalizada.objects.filter(user=request.user).first()
        mensagem = mensagem_obj.mensagem if mensagem_obj else ''
        mensagem_url = mensagem_obj.imagem.url if mensagem_obj and mensagem_obj.imagem else None

        # Obter perfil do usuário
        perfil = Perfil.objects.get(user=request.user)
        nome_primeiro_conjuge = perfil.nome_primeiro_conjuge
        nome_segundo_conjuge = perfil.nome_segundo_conjuge
        data_casamento = perfil.data_casamento

        # Adicionar o link do WhatsApp personalizado para cada convidado
        for convidado in convidados:  # Mudar de nao_confirmados para convidados
            telefone = convidado.whatsapp
            if telefone:
                if not telefone.startswith("55"):
                    telefone = f"55{telefone}"

                endereco_completo = f"Rua {perfil.rua}, {perfil.numero} - {perfil.bairro}, {perfil.municipio}/{perfil.estado} - {perfil.cep}"
                data_formatada = perfil.data_casamento.strftime("%d/%m/%Y")
                horario_formatado = perfil.horario_casamento.strftime("%H:%M") if perfil.horario_casamento else ""

                # Personalizar mensagem
                mensagem_formatada = mensagem.replace("{nome}", convidado.nome_convidado)\
                            .replace("{link}", convidado.link_convite)\
                            .replace("{local}", endereco_completo)\
                            .replace("{data}", data_formatada)\
                            .replace("{horario}", horario_formatado)\
                            .replace("{noivo}", perfil.nome_primeiro_conjuge)\
                            .replace("{noiva}", perfil.nome_segundo_conjuge)

                # Adicionar link da imagem, se existir
                if mensagem_obj and mensagem_obj.imagem:
                    imagem_url = request.build_absolute_uri(mensagem_obj.imagem.url)
                    mensagem_formatada += f"\n\n📷 Segue também um lindo convite que preparamos para você: {imagem_url}"

                mensagem_encoded = quote(mensagem_formatada)
                convidado.link_whatsapp = f"https://wa.me/{telefone}?text={mensagem_encoded}"
            else:
                convidado.link_whatsapp = "#"

        return render(request, 'lista_convidados.html', {
            'convidados': convidados, 
            'nao_confirmados': nao_confirmados,
            'mensagem': mensagem,
            'mensagem_url': mensagem_url,
            'perfil': perfil,
            'nome_primeiro_conjuge': nome_primeiro_conjuge,
            'nome_segundo_conjuge': nome_segundo_conjuge,
            'data_casamento': data_casamento
        })
    
    elif request.method == 'POST':
        nome_convidado = request.POST.get('nome_convidado')
        whatsapp = request.POST.get('whatsapp')
        maximo_acompanhantes = request.POST.get('maximo_acompanhantes', '0')  # Obtém o valor ou '0' se vazio
        maximo_acompanhantes = int(maximo_acompanhantes) if maximo_acompanhantes.isdigit() else 0  # Converte ou usa 0
        
        # Criar um novo convidado e salvar o link do WhatsApp automaticamente
        Convidados.objects.create(
            user=request.user,
            nome_convidado=nome_convidado,
            whatsapp=whatsapp,
            maximo_acompanhantes=maximo_acompanhantes
        )
        
        return redirect('lista_convidados')






def cadastrar_convidados_em_lote(request):
    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo_convidados')
        if not arquivo:
            messages.error(request, "Nenhum arquivo foi enviado.")
            return redirect('lista_convidados')

        try:
            # Detecta o tipo de arquivo (CSV ou Excel)
            if arquivo.name.endswith('.csv'):
                df = pd.read_csv(arquivo)
            elif arquivo.name.endswith('.xlsx'):
                df = pd.read_excel(arquivo)
            else:
                messages.error(request, "Formato de arquivo não suportado. Envie um arquivo CSV ou Excel.")
                return redirect('lista_convidados')

            # Verifica se as colunas necessárias estão presentes
            colunas_esperadas = ['Nome do convidado', 'Whatsapp', 'Máximo de Acompanhantes']
            if not all(coluna in df.columns for coluna in colunas_esperadas):
                messages.error(request, f"O arquivo deve conter as colunas: {', '.join(colunas_esperadas)}.")
                return redirect('lista_convidados')

            # Itera pelas linhas e cria os convidados
            for _, row in df.iterrows():
                Convidados.objects.create(
                    user=request.user,
                    nome_convidado=row['Nome do convidado'],
                    whatsapp=str(row['Whatsapp']),
                    maximo_acompanhantes=int(row['Máximo de Acompanhantes'])
                )

            messages.success(request, "Convidados cadastrados com sucesso!")
        except Exception as e:
            messages.error(request, f"Ocorreu um erro ao processar o arquivo: {str(e)}")
        return redirect('lista_convidados')


def exportar_convidados_excel(request):
    # Criar um novo arquivo Excel
    wb = openpyxl.Workbook()

    # Filtrar os convidados do usuário autenticado
    usuario_responsavel = request.user

    # Planilha para confirmados
    ws_confirmados = wb.create_sheet('Confirmados')
    # Adicionar os títulos das colunas
    ws_confirmados.append(['Convidado', 'Whatsapp', 'Acompanhante', 'Link', 'Status'])
    # Obter convidados confirmados do usuário responsável
    convidados_confirmados = Convidados.objects.filter(user=usuario_responsavel, status='C')

    for convidado in convidados_confirmados:
        for acompanhante in convidado.acompanhantes.all():
            # Adicionar uma linha para cada acompanhante
            ws_confirmados.append([
                convidado.nome_convidado, 
                convidado.whatsapp, 
                acompanhante.nome, 
                convidado.link_convite, 
                convidado.get_status_display()
            ])

    # Planilha para aguardando confirmação
    ws_aguardando = wb.create_sheet('Aguardando confirmação')
    # Adicionar os títulos das colunas
    ws_aguardando.append(['Convidado', 'Whatsapp', 'Máximo de acompanhantes', 'Link', 'Status'])
    # Obter convidados aguardando confirmação do usuário responsável
    convidados_aguardando = Convidados.objects.filter(user=usuario_responsavel, status='AC')

    for convidado in convidados_aguardando:
        ws_aguardando.append([
            convidado.nome_convidado, 
            convidado.whatsapp, 
            convidado.maximo_acompanhantes, 
            convidado.link_convite, 
            convidado.get_status_display()
        ])

    # Planilha para total de confirmados (incluindo acompanhantes)
    ws_total_confirmados = wb.create_sheet('Total Confirmados')
    # Adicionar os títulos das colunas
    ws_total_confirmados.append(['Nome', 'Whatsapp', 'Tipo'])
    # Inicializar contador para o total de pessoas
    total_pessoas = 0

    for convidado in convidados_confirmados:
        # Adicionar o convidado à lista
        ws_total_confirmados.append([convidado.nome_convidado, convidado.whatsapp, 'Convidado'])
        total_pessoas += 1

        # Adicionar os acompanhantes do convidado à lista
        for acompanhante in convidado.acompanhantes.all():
            ws_total_confirmados.append([acompanhante.nome, '', 'Acompanhante'])  # Deixe o WhatsApp vazio para acompanhantes
            total_pessoas += 1

    # Adicionar o total de pessoas no final da planilha
    ws_total_confirmados.append([])
    ws_total_confirmados.append(['Total de pessoas:', '', total_pessoas])

    # Remover a planilha padrão que é criada automaticamente
    del wb['Sheet']

    # Definir a resposta HTTP para o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=convidados.xlsx'

    # Salvar o arquivo Excel na resposta HTTP
    wb.save(response)
    return response




def excluir_presente(request, presente_id):
    presente = get_object_or_404(Presentes, id=presente_id)
    presente.delete()
    return redirect('home')

def excluir_convidado(request, convidado_id):
    # Recuperar o convidado pelo ID
    convidado = get_object_or_404(Convidados, id=convidado_id)

    # Verificar se o convidado tem algum presente reservado
    presentes_reservados = Presentes.objects.filter(reservado_por=convidado)

    if presentes_reservados.exists():
        # Tornar o presente reservado disponível novamente
        presente = presentes_reservados.first()  # Aqui pegamos o primeiro presente reservado
        presente.reservado_por = None  # Remove a associação com o convidado
        presente.reservado = False  # Também pode definir como não reservado
        presente.save()

    # Excluir o convidado
    convidado.delete()

    # Mensagem de sucesso
    messages.success(request, "O convidado foi excluído e o presente reservado está disponível novamente.")
    return redirect('lista_convidados')


def enviar_mensagens(request):
    if request.method == "POST":
        data = request.POST
        convidados_ids = json.loads(data.get('convidados', []))
        convidados = Convidados.objects.filter(id__in=convidados_ids)

        # Obter a mensagem personalizada
        try:
            mensagem_personalizada      = MensagemPersonalizada.objects.filter(user=request.user).latest('data_criacao').mensagem
            mensagem_personalizada_obj  = MensagemPersonalizada.objects.filter(user=request.user).latest('data_criacao')
            arquivo                     = mensagem_personalizada_obj.imagem  # Recuperar o arquivo do banco

        except MensagemPersonalizada.DoesNotExist:
            mensagem_personalizada = ""
            return render(request, 'lista_convidados.html', {'mensagem_personalizada': mensagem_personalizada})

        # Verificar se o arquivo é válido (baseado na extensão)
        if arquivo and not arquivo.name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            raise ValidationError('O arquivo enviado não é uma imagem válida.')

        # Salvar a imagem temporariamente no disco C:
        temp_dir = r'C:\Temp'
        os.makedirs(temp_dir, exist_ok=True)  # Garantir que o diretório existe
        arquivo_temporario = None

        if arquivo:
            arquivo_nome        = os.path.basename(arquivo.name)
            arquivo_temporario  = os.path.join(temp_dir, arquivo_nome)

            with open(arquivo_temporario, 'wb') as temp_file:

                for chunk in arquivo.chunks():
                    temp_file.write(chunk)

        # Abrir WhatsApp Web
        webbrowser.open('https://web.whatsapp.com/')
        print("Aguardando o WhatsApp Web carregar...")
        sleep(30)

            # Enviar o arquivo, se existir
        if arquivo_temporario:
            pyautogui.hotkey('ctrl', 'o')  # Atalho para abrir upload de arquivo
            sleep(2)
            pyautogui.write(arquivo_temporario)  # Caminho do arquivo
            sleep(2)
            pyautogui.press('enter')  # Confirmar upload
            sleep(2)
            pyautogui.hotkey('ctrl', 'c')
            sleep(2)
            pyautogui.hotkey('alt', 'left')
            sleep(15)

        # Lista para armazenar os convidados que falharam
        erros_envio = []

        for convidado in convidados:
            nome        = convidado.nome_convidado
            telefone    = convidado.whatsapp

            if not telefone.startswith("55"):
                telefone = f"55{telefone}"

            link        = convidado.link_convite
            mensagem    = mensagem_personalizada.replace("{nome}", nome).replace("{link}", link)

            # Enviar a mensagem de texto
            link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
            

            try:

                
                webbrowser.open(link_mensagem_whatsapp)
                print(f"Enviando mensagem para {nome}...")
                sleep(15)
                pyautogui.press('enter')  # Enviar mensagem
                sleep(5)
                pyautogui.hotkey('ctrl', 'v')
                sleep(2)
                pyautogui.press('enter')
                sleep(5)

                pyautogui.hotkey('ctrl', 'w')  # Fechar a guia após o envio
                sleep(10)
            except Exception as e:
                print(f"Erro ao enviar mensagem para {nome}: {e}")
                erros_envio.append({'nome': nome, 'telefone': telefone, 'link': link})

        # Remover arquivo temporário, se foi criado
        if arquivo_temporario and os.path.exists(arquivo_temporario):
            os.remove(arquivo_temporario)

        # Se houver erros, gerar o arquivo .csv para download
        if erros_envio:
            # Criar o CSV na memória
            buffer = StringIO()
            escritor_csv = csv.DictWriter(buffer, fieldnames=['nome', 'telefone', 'link'])
            escritor_csv.writeheader()
            escritor_csv.writerows(erros_envio)

            # Configurar a resposta HTTP para o download do arquivo
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="erros_envio.csv"'
            return response

        # Caso não haja erros, retornar sucesso
        return JsonResponse({'status': 'sucesso'})

    


@csrf_exempt
def salvar_mensagem(request):
    if request.method == 'POST':
        try:
            mensagem = request.POST.get('mensagem', '')
            arquivo = request.FILES.get('arquivo', None)  # Obtém o arquivo enviado
            user = request.user  # Obtendo o usuário logado

            # Verificar se já existe uma mensagem para o usuário
            mensagem_existente = MensagemPersonalizada.objects.filter(user=user).first()

            if mensagem_existente:
                # Atualizar a mensagem existente
                mensagem_existente.mensagem = mensagem
                if arquivo:
                    # Salvar o novo arquivo e substituir o antigo
                    mensagem_existente.imagem = arquivo
                mensagem_existente.save()
            else:
                # Criar uma nova mensagem se não existir
                nova_mensagem = MensagemPersonalizada(user=user, mensagem=mensagem)
                if arquivo:
                    # Salvar a imagem no banco de dados
                    nova_mensagem.imagem = arquivo
                nova_mensagem.save()

            # Retorna a mensagem salva ou atualizada para o frontend
            return JsonResponse({
                'success': True,
                'mensagem': mensagem  # Retorna a mensagem salva para exibição
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
        



import requests
from django.core.files.base import ContentFile
from io import BytesIO
from PIL import Image

def buscar_detalhes_produto(request):
    url = request.GET.get("url", "")
    if not url:
        return JsonResponse({"error": "Nenhuma URL fornecida."}, status=400)

    url_decodificada = unquote(url)
    print(f"URL Decodificada: {url_decodificada}")
    
    # Ajustando a expressão regular para capturar item_id, wid e também IDs com hífen
    match = re.search(r'(item_id:MLB\d+|wid=MLB\d+|MLB-\d+)', url_decodificada)

    if not match:
        return JsonResponse({"error": "ID do produto não encontrado no link."}, status=400)

    # Verificando qual ID foi encontrado (item_id, wid ou o ID com hífen)
    produto_id = match.group(0).split(":")[-1] if "item_id" in match.group(0) else match.group(0).split("=")[-1]

    # Removendo o hífen, caso presente
    produto_id = produto_id.replace('-', '')
    
    print(f"ID do produto extraído: {produto_id}")

    try:
        api_url = f"https://api.mercadolibre.com/items/{produto_id}"
        print(f"link completo da api: {api_url}")
        headers = {
            "Authorization": "Bearer APP_USR-3067363791536171-040319-9d7c887e455b5546399871bdb202457f-153067470",
            "Content-Type": "application/json"
        }

        response = requests.get(api_url, headers=headers)
        
        if response.status_code == 200:
            data = response.json()

            pictures = data.get("pictures", [])
            imagem_url = pictures[0].get("secure_url", "") if pictures else ""

            img_file = None
            if imagem_url:
                img_response = requests.get(imagem_url)
                if img_response.status_code == 200:
                    img = Image.open(BytesIO(img_response.content))
                    img_name = f"produto_{produto_id}.jpg"
                    img_io = BytesIO()
                    img.save(img_io, format='JPEG')
                    img_file = ContentFile(img_io.getvalue(), name=img_name)

            # Verifica se já existe um presente com esse nome para evitar duplicidade
            nome_presente = data.get("title", "")
            if not Presentes.objects.filter(nome_presente=nome_presente, user=request.user).exists():
                Presentes.objects.create(
                    user=request.user,
                    nome_presente=nome_presente,
                    foto=img_file,
                    preco=data.get("price", 0),
                    importancia=0,
                    link_sugestao_compra=data.get("permalink", ""),
                )

            return JsonResponse({
                "nome": nome_presente,
                "preco": f"{data.get('price', '0,00')}",
                "imagem": imagem_url,
            })
        else:
            return JsonResponse({"error": "Produto não encontrado ou erro na API."}, status=404)

    except Exception as e:
        return JsonResponse({"error": f"Erro ao buscar o produto: {str(e)}"}, status=500)


@login_required(login_url='/auth/logar/')
def admin_produto(request):
    if not request.user.is_staff:
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('home')

    if request.method == 'POST':
        nome = request.POST.get('nome_produto')
        preco = request.POST.get('preco')
        link_sugestao = request.POST.get('link_sugestao_compra')
        link_cobranca = request.POST.get('link_cobranca')
        descricao = request.POST.get('descricao', '')
        categoria = request.POST.get('categoria', '')
        importancia = request.POST.get('importancia', 3)
        foto = request.FILES.get('foto')

        if ',' in preco:
            preco = preco.replace(',', '.')

        produto = ProdutoBase.objects.create(
            nome=nome,
            preco=preco,
            link_sugestao_compra=link_sugestao,
            link_cobranca=link_cobranca,
            descricao=descricao,
            categoria=categoria,
            importancia=importancia,
            foto=foto
        )

        messages.success(request, 'Produto adicionado com sucesso!')
        return redirect('admin_produto')

    # Mostrar todos os produtos, não apenas os ativos
    produtos = ProdutoBase.objects.all().order_by('-data_criacao')
    return render(request, 'admin_produto.html', {'produtos': produtos})


@login_required(login_url='/auth/logar/')
def selecionar_presentes_prontos(request):
    produtos_base = ProdutoBase.objects.filter(ativo=True)
    return render(request, 'selecionar_presentes.html', {'produtos_base': produtos_base})

@login_required(login_url='/auth/logar/')
def importar_lista_padrao(request):
    if request.method == 'POST':
        produtos_selecionados = request.POST.getlist('produtos_selecionados')
        
        for produto_id in produtos_selecionados:
            produto = ProdutoBase.objects.get(id=produto_id)
            Presentes.objects.get_or_create(
                user=request.user,
                produto_base=produto,
                defaults={
                    'nome_presente': produto.nome,
                    'foto': produto.foto,
                    'preco': produto.preco,
                    'importancia': produto.importancia,
                    'link_sugestao_compra': produto.link_sugestao_compra,
                    'link_cobranca': produto.link_cobranca
                }
            )
        
        messages.success(request, 'Presentes selecionados foram adicionados com sucesso!')
        return redirect('home')
    
    return redirect('selecionar_presentes_prontos')

@login_required(login_url='/auth/logar/')
def editar_produto_base(request, produto_id):
    if not request.user.is_staff:
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('home')

    produto = get_object_or_404(ProdutoBase, id=produto_id)

    if request.method == 'POST':
        produto.nome = request.POST.get('nome_produto')
        produto.preco = request.POST.get('preco').replace(',', '.')
        produto.link_sugestao_compra = request.POST.get('link_sugestao_compra')
        produto.link_cobranca = request.POST.get('link_cobranca')
        produto.descricao = request.POST.get('descricao', '')
        produto.categoria = request.POST.get('categoria', '')
        produto.importancia = request.POST.get('importancia', 3)

        if 'foto' in request.FILES:
            produto.foto = request.FILES['foto']

        try:
            produto.save()
            messages.success(request, 'Produto atualizado com sucesso!')
            return redirect('admin_produto')
        except Exception as e:
            messages.error(request, f'Erro ao atualizar produto: {str(e)}')

    return render(request, 'editar_produto_base.html', {'produto': produto})


@login_required(login_url='/auth/logar/')
def desativar_produto_base(request, produto_id):
    if not request.user.is_staff:
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('home')

    produto = get_object_or_404(ProdutoBase, id=produto_id)
    produto.ativo = False
    produto.save()
    messages.success(request, 'Produto desativado com sucesso!')
    return redirect('admin_produto')

@login_required(login_url='/auth/logar/')
def ativar_produto_base(request, produto_id):
    if not request.user.is_staff:
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('home')

    produto = get_object_or_404(ProdutoBase, id=produto_id)
    produto.ativo = True
    produto.save()
    messages.success(request, 'Produto ativado com sucesso!')
    return redirect('admin_produto')

@login_required(login_url='/auth/logar/')
def excluir_produto_base(request, produto_id):
    if not request.user.is_staff:
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('home')

    produto = get_object_or_404(ProdutoBase, id=produto_id)
    
    try:
        produto.delete()
        messages.success(request, 'Produto excluído permanentemente com sucesso!')
    except Exception as e:
        messages.error(request, f'Erro ao excluir produto: {str(e)}')
    
    return redirect('admin_produto')